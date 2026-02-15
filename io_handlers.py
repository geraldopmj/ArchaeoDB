import os
import uuid
import html
import xlsxwriter
import openpyxl
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.platypus import Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
from database import Database

def resolve_fk(val, lookup_map, parent_id=None):
    """Auxiliar para resolver Nomes para IDs (Lookup reverso)."""
    if val is None or val == "": return None
    if isinstance(val, (int, float)): return int(val)
    val_str = str(val).strip()
    if val_str.isdigit(): return int(val_str)
    
    if parent_id is not None:
        return lookup_map.get((parent_id, val_str))
    return lookup_map.get(val_str)

def export_pdf(db, file_path, columns, visible_ids=None):
    """Gera um PDF com todos os materiais."""
    try:
        if visible_ids is not None:
            if not visible_ids:
                return False, "Nenhum material visível para exportar."
            placeholders = ','.join(['?'] * len(visible_ids))
            query = f"SELECT * FROM Material WHERE id IN ({placeholders})"
            cursor = db.conn.execute(query, visible_ids)
            materials = cursor.fetchall()
        else:
            materials = db.fetch("Material")

        if not materials:
            return False, "Nenhum material para exportar."

        # Mapas para substituir IDs por Nomes
        units = db.fetch("ExcavationUnit", "id, name")
        unit_map = {u[0]: u[1] for u in units}
        
        levels = db.fetch("Level", "id, level")
        level_map = {l[0]: l[1] for l in levels}
        
        # Mapa para Sítios (Unit ID -> Site Name) para preencher a coluna "Sítio"
        query_sites = """
            SELECT u.id, s.name 
            FROM ExcavationUnit u
            JOIN Assemblage a ON u.assemblage_id = a.id
            JOIN Collection c ON a.collection_id = c.id
            JOIN Site s ON c.site_id = s.id
        """
        cursor = db.conn.execute(query_sites)
        unit_site_map = {row[0]: row[1] for row in cursor.fetchall()}

        pdf = canvas.Canvas(file_path, pagesize=A4)
        largura, altura = A4
        
        styles = getSampleStyleSheet()
        style_normal = styles["BodyText"]
        style_normal.fontName = "Helvetica"
        style_normal.fontSize = 10

        for i, material in enumerate(materials):
            titulo = f"Material ID: {material[0]}"
            pdf.setFont("Helvetica-Bold", 16)
            pdf.drawCentredString(largura / 2, altura - 50, titulo)

            # Converte para lista para modificar
            material_list = list(material)
            
            # Insere o Nome do Sítio na posição 1 (para alinhar com as colunas da UI)
            unit_id = material_list[1]
            material_list.insert(1, unit_site_map.get(unit_id, ""))
            
            # Substitui IDs por Nomes (Unidade e Nível)
            if material_list[2] in unit_map:
                material_list[2] = unit_map[material_list[2]]
            if material_list[3] in level_map:
                material_list[3] = level_map[material_list[3]]

            dados_tabela = [["Campo", "Valor"]]
            limit = min(len(columns), len(material_list))
            for j in range(limit):
                val_str = str(material_list[j]) if material_list[j] is not None else ""
                p_val = Paragraph(html.escape(val_str), style_normal)
                dados_tabela.append([columns[j], p_val])

            tabela = Table(dados_tabela, colWidths=[150, 350])
            estilo = TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica'),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ])
            tabela.setStyle(estilo)
            tabela.wrapOn(pdf, largura, altura)
            tabela.drawOn(pdf, 50, altura - 25 - len(dados_tabela) * 20)

            if i < len(materials) - 1:
                pdf.showPage()

        pdf.save()
        return True, "PDF gerado com sucesso!"
    except Exception as e:
        return False, f"Erro ao gerar PDF: {str(e)}"

def export_units_pdf(db, file_path, visible_ids=None):
    """Gera um PDF com todas as unidades de escavação e seus níveis."""
    try:
        if visible_ids is not None:
            if not visible_ids:
                return False, "Nenhuma unidade visível para exportar."
            placeholders = ','.join(['?'] * len(visible_ids))
            query = f"SELECT * FROM ExcavationUnit WHERE id IN ({placeholders})"
            cursor = db.conn.execute(query, visible_ids)
            units = cursor.fetchall()
        else:
            units = db.fetch("ExcavationUnit")

        if not units:
            return False, "Nenhuma unidade para exportar."

        pdf = canvas.Canvas(file_path, pagesize=A4)
        width, height = A4
        y = height - 50

        styles = getSampleStyleSheet()
        style_cell = styles["BodyText"]
        style_cell.fontName = "Helvetica"
        style_cell.fontSize = 9

        for unit in units:
            # Verifica espaço para o cabeçalho da unidade (aprox 150pts de margem de segurança)
            if y < 150:
                pdf.showPage()
                y = height - 50

            # Recupera nome da Amostra para contexto
            assemblage = db.fetch("Assemblage", "name", {"id": unit[1]})
            assemblage_name = assemblage[0][0] if assemblage else "Desconhecida"

            # Contagem de materiais
            materials = db.fetch("Material", "id", {"excavation_unit_id": unit[0]})
            material_count = len(materials)

            # Título da Unidade
            pdf.setFont("Helvetica-Bold", 14)
            pdf.setFillColor(colors.black)
            title = f"Unidade: {unit[2]} (Amostra: {assemblage_name})"
            pdf.drawString(50, y, title)
            y -= 20
            
            # Detalhes da Unidade
            pdf.setFont("Helvetica", 10)
            details = [
                f"Tipo: {unit[3] or ''} | Tamanho: {unit[4] or ''}",
                f"Lat: {unit[5] or ''} | Lon: {unit[6] or ''}",
                f"Data: {unit[14] or ''} a {unit[15] or ''}",
                f"Responsável: {unit[13] or ''} | Obs: {unit[17] or ''}",
                f"Total de Vestígios: {material_count}"
            ]
            
            for line in details:
                pdf.drawString(50, y, line)
                y -= 15
            
            y -= 10 # Espaço antes da tabela de níveis

            # Busca Níveis da Unidade
            levels = db.fetch("Level", "*", {"excavation_unit_id": unit[0]})
            
            if levels:
                pdf.setFont("Helvetica-Bold", 11)
                pdf.drawString(50, y, "Níveis:")
                y -= 20
                
                # Cabeçalho da Tabela de Níveis
                data = [["Nível", "Prof. Ini", "Prof. Fin", "Cor", "Textura", "Descrição"]]
                for lvl in levels:
                    row = []
                    for item in [lvl[2], lvl[3], lvl[4], lvl[5], lvl[6], lvl[7]]:
                        val_str = str(item or "")
                        row.append(Paragraph(html.escape(val_str), style_cell))
                    data.append(row)
                
                col_widths = [50, 60, 60, 80, 80, 160]
                t = Table(data, colWidths=col_widths)
                style = TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 9),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.whitesmoke),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ])
                t.setStyle(style)
                
                # Calcula altura da tabela
                w, h = t.wrap(width - 100, height)
                
                # Se a tabela não couber na página, quebra página
                if y - h < 50:
                    pdf.showPage()
                    y = height - 50
                    pdf.setFont("Helvetica-Bold", 11)
                    pdf.drawString(50, y, f"Níveis (Unidade: {unit[2]} - Cont.):")
                    y -= 20
                
                t.drawOn(pdf, 50, y - h)
                y -= (h + 20)
            else:
                pdf.setFont("Helvetica-Oblique", 10)
                pdf.drawString(50, y, "Nenhum nível registrado.")
                y -= 20

            # Linha separadora entre unidades
            pdf.setStrokeColor(colors.lightgrey)
            pdf.line(50, y, width - 50, y)
            y -= 20

        pdf.save()
        return True, "PDF de Unidades gerado com sucesso!"
    except Exception as e:
        return False, f"Erro ao gerar PDF: {str(e)}"

def export_xlsx(db, file_path):
    """Exporta todo o banco de dados para Excel."""
    try:
        workbook = xlsxwriter.Workbook(file_path)
        tabelas = ['Site', 'Collection', 'Assemblage', 'ExcavationUnit', 'Level', 'Material']

        # Mapas para substituir IDs por Nomes
        sites = db.fetch("Site", "id, name")
        site_map = {s[0]: s[1] for s in sites}
        
        cols = db.fetch("Collection", "id, name")
        col_map = {c[0]: c[1] for c in cols}
        
        asms = db.fetch("Assemblage", "id, name")
        asm_map = {a[0]: a[1] for a in asms}
        
        units = db.fetch("ExcavationUnit", "id, name")
        unit_map = {u[0]: u[1] for u in units}
        
        levels = db.fetch("Level", "id, level")
        level_map = {l[0]: l[1] for l in levels}

        for tabela in tabelas:
            worksheet = workbook.add_worksheet(tabela)
            dados = db.fetch(tabela)
            
            cursor = db.conn.execute(f"SELECT * FROM {tabela} LIMIT 0")
            colunas = [desc[0] for desc in cursor.description]
            
            for col_num, coluna in enumerate(colunas):
                worksheet.write(0, col_num, coluna)
            
            for row_num, linha in enumerate(dados, start=1):
                linha_list = list(linha)
                
                # Substituições de FKs por Nomes
                if tabela == 'Collection':
                    # site_id é coluna 1
                    if len(linha_list) > 1 and linha_list[1] in site_map:
                        linha_list[1] = site_map[linha_list[1]]
                elif tabela == 'Assemblage':
                    # collection_id é coluna 1
                    if len(linha_list) > 1 and linha_list[1] in col_map:
                        linha_list[1] = col_map[linha_list[1]]
                elif tabela == 'ExcavationUnit':
                    # assemblage_id é coluna 1
                    if len(linha_list) > 1 and linha_list[1] in asm_map:
                        linha_list[1] = asm_map[linha_list[1]]
                elif tabela == 'Level':
                    # excavation_unit_id é coluna 1
                    if len(linha_list) > 1 and linha_list[1] in unit_map:
                        linha_list[1] = unit_map[linha_list[1]]
                elif tabela == 'Material':
                    # excavation_unit_id é coluna 1
                    if len(linha_list) > 1 and linha_list[1] in unit_map:
                        linha_list[1] = unit_map[linha_list[1]]
                    # level_id é coluna 2
                    if len(linha_list) > 2 and linha_list[2] in level_map:
                        linha_list[2] = level_map[linha_list[2]]

                for col_num, valor in enumerate(linha_list):
                    worksheet.write(row_num, col_num, valor)

        workbook.close()
        return True, "Arquivo Excel gerado com sucesso!"
    except Exception as e:
        return False, f"Erro ao exportar Excel: {str(e)}"

def update_database_from_excel(db, xlsx_path):
    """Atualiza o banco de dados atual a partir de um arquivo Excel."""
    try:
        wb = openpyxl.load_workbook(xlsx_path)
        # Ordem: Site -> Collection -> Assemblage -> ExcavationUnit -> Level -> Material
        tables = ['Site', 'Collection', 'Assemblage', 'ExcavationUnit', 'Level', 'Material']
        
        site_map = {}
        col_map = {}
        asm_map = {}
        unit_map = {}
        level_map = {}
        
        for table in tables:
            if table in wb.sheetnames:
                ws = wb[table]
                rows = list(ws.iter_rows(values_only=True))
                
                if not rows:
                    continue
                
                # Atualiza mapas de lookup baseados nos dados atuais do banco
                if table == 'Collection':
                    site_map = {str(n): i for n, i in db.fetch("Site", "name, id")}
                elif table == 'Assemblage':
                    col_map = {str(n): i for n, i in db.fetch("Collection", "name, id")}
                elif table == 'ExcavationUnit':
                    asm_map = {str(n): i for n, i in db.fetch("Assemblage", "name, id")}
                elif table == 'Level':
                    unit_map = {str(n): i for n, i in db.fetch("ExcavationUnit", "name, id")}
                elif table == 'Material':
                    unit_map = {str(n): i for n, i in db.fetch("ExcavationUnit", "name, id")}
                    level_data = db.fetch("Level", "id, excavation_unit_id, level")
                    level_map = {(u, str(l)): i for i, u, l in level_data}
                    
                headers = rows[0]
                # Normaliza os nomes das colunas: str, lower, strip, spaces->underscore
                def norm(h):
                    if h is None:
                        return None
                    return str(h).strip().lower().replace(' ', '_')

                headers = [norm(h) for h in headers]
                headers = [h for h in headers if h]
                # Mapeamento de variantes em PT/EN para os nomes das colunas do DB
                alias_map = {
                    'amostra': 'assemblage_id', 'amostra_id': 'assemblage_id', 'assemblage_id': 'assemblage_id',
                    'unidade_de_escavacao': 'excavation_unit_id', 'unidade_escavacao': 'excavation_unit_id',
                    'excavation_unit_id': 'excavation_unit_id', 'id_unidade': 'excavation_unit_id',
                    'sitio': 'site_id', 'colecao': 'collection_id', 'nivel': 'level_id'
                }
                mapped_headers = []
                for h in headers:
                    mapped_headers.append(alias_map.get(h, h))
                
                for i, row_values in enumerate(rows[1:], start=2):
                    try:
                        # Converte valores e aplica mapeamento de headers
                        values = list(row_values)
                        # Ajusta o comprimento se row shorter/longer
                        if len(values) < len(mapped_headers):
                            values += [None] * (len(mapped_headers) - len(values))
                        data = {}
                        for key, val in zip(mapped_headers, values):
                            if val is None:
                                data[key] = None
                            else:
                                # Tenta converter IDs para int para evitar datatype mismatch
                                if key == 'id' or key.endswith('_id'):
                                    try:
                                        # Tenta converter para float primeiro (para lidar com "1.0") depois int
                                        data[key] = int(float(val))
                                    except Exception:
                                        data[key] = val
                                else:
                                    if isinstance(val, str):
                                        data[key] = val.strip()
                                    else:
                                        data[key] = val
                        
                        # Resolve FKs (Nome -> ID)
                        if table == 'Collection':
                            val = data.get('site_id')
                            resolved = resolve_fk(val, site_map)
                            if resolved is None and val is not None:
                                return False, f"Erro na tabela '{table}', linha {i}: Sítio '{val}' não encontrado."
                            data['site_id'] = resolved
                        elif table == 'Assemblage':
                            val = data.get('collection_id')
                            resolved = resolve_fk(val, col_map)
                            if resolved is None and val is not None:
                                return False, f"Erro na tabela '{table}', linha {i}: Coleção '{val}' não encontrada."
                            data['collection_id'] = resolved
                        elif table == 'ExcavationUnit':
                            val = data.get('assemblage_id')
                            resolved = resolve_fk(val, asm_map)
                            if resolved is None and val is not None:
                                return False, f"Erro na tabela '{table}', linha {i}: Amostra '{val}' não encontrada."
                            data['assemblage_id'] = resolved
                        elif table == 'Level':
                            val = data.get('excavation_unit_id')
                            resolved = resolve_fk(val, unit_map)
                            if resolved is None and val is not None:
                                return False, f"Erro na tabela '{table}', linha {i}: Unidade '{val}' não encontrada."
                            data['excavation_unit_id'] = resolved
                        elif table == 'Material':
                            val = data.get('excavation_unit_id')
                            resolved = resolve_fk(val, unit_map)
                            if resolved is None and val is not None:
                                return False, f"Erro na tabela '{table}', linha {i}: Unidade '{val}' não encontrada."
                            data['excavation_unit_id'] = resolved
                            data['level_id'] = resolve_fk(data.get('level_id'), level_map, parent_id=data.get('excavation_unit_id'))

                        if table == 'Material':
                            # Remove campos obsoletos se existirem no Excel (se houver)
                            data.pop('provenience', None)
                            data.pop('strata', None)

                            record_uuid = data.get('uuid')
                            if record_uuid:
                                exists = db.fetch(table, "id", {"uuid": record_uuid})
                                if exists:
                                    update_data = {k: v for k, v in data.items() if k != 'id' and k != 'uuid'}
                                    if update_data:
                                        db.update(table, update_data, {"uuid": record_uuid})
                                else:
                                    insert_data = {k: v for k, v in data.items() if k != 'id'}
                                    db.insert(table, insert_data)
                            else:
                                data['uuid'] = str(uuid.uuid4())
                                insert_data = {k: v for k, v in data.items() if k != 'id'}
                                db.insert(table, insert_data)
                        else:
                            record_id = data.get('id')
                            if record_id is not None:
                                exists = db.fetch(table, "id", {"id": record_id})
                                if exists:
                                    update_data = {k: v for k, v in data.items() if k != 'id'}
                                    if update_data:
                                        db.update(table, update_data, {"id": record_id})
                                else:
                                    db.insert(table, data)
                            else:
                                db.insert(table, data)
                    except Exception as e:
                        return False, f"Erro na tabela '{table}', linha {i}: {str(e)}"

        return True, "Banco de dados atualizado com sucesso!"
    except Exception as e:
        return False, f"Erro ao atualizar do Excel: {str(e)}"

def import_database_from_excel(xlsx_path, db_path):
    """Cria ou atualiza um banco de dados a partir de um arquivo Excel."""
    try:
        temp_db = Database(db_path, create_tables=True)
        wb = openpyxl.load_workbook(xlsx_path)
        # Ordem: Site -> Collection -> Assemblage -> ExcavationUnit -> Level -> Material
        tables = ['Site', 'Collection', 'Assemblage', 'ExcavationUnit', 'Level', 'Material']
        
        site_map = {}
        col_map = {}
        asm_map = {}
        unit_map = {}
        level_map = {}
        
        for table in tables:
            if table in wb.sheetnames:
                ws = wb[table]
                rows = list(ws.iter_rows(values_only=True))
                
                if not rows:
                    continue
                
                # Atualiza mapas de lookup baseados nos dados já inseridos no temp_db
                if table == 'Collection':
                    site_map = {str(n): i for n, i in temp_db.fetch("Site", "name, id")}
                elif table == 'Assemblage':
                    col_map = {str(n): i for n, i in temp_db.fetch("Collection", "name, id")}
                elif table == 'ExcavationUnit':
                    asm_map = {str(n): i for n, i in temp_db.fetch("Assemblage", "name, id")}
                elif table == 'Level' or table == 'Material':
                    unit_map = {str(n): i for n, i in temp_db.fetch("ExcavationUnit", "name, id")}
                    level_data = temp_db.fetch("Level", "id, excavation_unit_id, level")
                    level_map = {(u, str(l)): i for i, u, l in level_data}
                    
                headers = rows[0]
                def norm(h):
                    if h is None:
                        return None
                    return str(h).strip().lower().replace(' ', '_')

                headers = [norm(h) for h in headers]
                headers = [h for h in headers if h]
                alias_map = {
                    'amostra': 'assemblage_id', 'amostra_id': 'assemblage_id', 'assemblage_id': 'assemblage_id',
                    'unidade_de_escavacao': 'excavation_unit_id', 'unidade_escavacao': 'excavation_unit_id',
                    'excavation_unit_id': 'excavation_unit_id', 'id_unidade': 'excavation_unit_id'
                }
                mapped_headers = [alias_map.get(h, h) for h in headers]

                for i, row_values in enumerate(rows[1:], start=2):
                    values = list(row_values)
                    if len(values) < len(mapped_headers):
                        values += [None] * (len(mapped_headers) - len(values))
                    data = {}
                    for key, val in zip(mapped_headers, values):
                        if key is None: continue
                        if val is None:
                            data[key] = None
                        else:
                            if key.endswith('_id') and isinstance(val, float):
                                try:
                                    data[key] = int(val)
                                except Exception:
                                    data[key] = val
                            else:
                                    if isinstance(val, str):
                                        data[key] = val.strip()
                                    else:
                                        data[key] = val

                        # Resolve FKs (Nome -> ID)
                        if table == 'Collection':
                            val = data.get('site_id')
                            resolved = resolve_fk(val, site_map)
                            if resolved is None and val is not None:
                                return False, f"Erro na tabela '{table}', linha {i}: Sítio '{val}' não encontrado."
                            data['site_id'] = resolved
                        elif table == 'Assemblage':
                            val = data.get('collection_id')
                            resolved = resolve_fk(val, col_map)
                            if resolved is None and val is not None:
                                return False, f"Erro na tabela '{table}', linha {i}: Coleção '{val}' não encontrada."
                            data['collection_id'] = resolved
                        elif table == 'ExcavationUnit':
                            val = data.get('assemblage_id')
                            resolved = resolve_fk(val, asm_map)
                            if resolved is None and val is not None:
                                return False, f"Erro na tabela '{table}', linha {i}: Amostra '{val}' não encontrada."
                            data['assemblage_id'] = resolved
                        elif table == 'Level':
                            val = data.get('excavation_unit_id')
                            resolved = resolve_fk(val, unit_map)
                            if resolved is None and val is not None:
                                return False, f"Erro na tabela '{table}', linha {i}: Unidade '{val}' não encontrada."
                            data['excavation_unit_id'] = resolved
                        elif table == 'Material':
                            val = data.get('excavation_unit_id')
                            resolved = resolve_fk(val, unit_map)
                            if resolved is None and val is not None:
                                return False, f"Erro na tabela '{table}', linha {i}: Unidade '{val}' não encontrada."
                            data['excavation_unit_id'] = resolved
                            data['level_id'] = resolve_fk(data.get('level_id'), level_map, parent_id=data.get('excavation_unit_id'))

                    if table == 'Material':
                        # Remove campos obsoletos se existirem no Excel
                        data.pop('provenience', None)
                        data.pop('strata', None)
                        
                        if 'uuid' not in data or not data['uuid']:
                            data['uuid'] = str(uuid.uuid4())

                    try:
                        record_id = data.get('id')
                        if record_id is not None:
                            exists = temp_db.fetch(table, "id", {"id": record_id})
                            if exists:
                                update_data = {k: v for k, v in data.items() if k != 'id'}
                                if update_data:
                                    temp_db.update(table, update_data, {"id": record_id})
                            else:
                                temp_db.insert(table, data)
                        else:
                            temp_db.insert(table, data)
                    except Exception as e:
                        return False, f"Erro na tabela '{table}', linha {i}: {str(e)}"
                        
        temp_db.close()
        return True, "Importação concluída com sucesso!"
    except Exception as e:
        return False, f"Erro ao importar do Excel: {str(e)}"
